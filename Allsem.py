import requests as req
import certifi
import urllib3
from requests.exceptions import MissingSchema

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import webbrowser

from bs4 import BeautifulSoup

import tkinter as tk
from tkinter import *
from tkinter import ttk, messagebox

import os

valueTable_CSE = {'120':'ENGLISH LANGUAGE LAB', '121':'C PROGRAMMING LAB', '171':'MATHEMATICS-I', '173':'CHEMISTRY', '175':'PROGRAMMING FOR PROBLEM SOLVING', '185':'ENVIRONMENTAL SCIENCE', '178':'CHEMISTRY LAB', '180':'PROGRAMMING FOR PROBLEM SOLVING LAB', '182':'WORKSHOP LAB', '176':'ESSENCE OF INDIAN TRADITIONAL KNOWLEDGE', '271':'ENGLISH', '272':'MATHEMATICS-II', '275':'PHYSICS', '276':'BASIC ELECTRICAL ENGINEERING', '287A':'ENGINEERING GRAPHICS LAB', '281':'ENGLISH LAB', '285':'PHYSICS LAB', '286':'BASIC ELECTRICAL ENGINEERING LAB', '279':'INDIAN CONSTITUTION', '301E':'OOP USING JAVA', '301F':'OOP USING JAVA LAB', '318':'DISCRETE MATHEMATICS', '358':'BASIC ELECTRONICS LAB', '383':'DATA STRUCTURES AND ALGORITHMS', '370':'DATA STRUC.AND ALGORITHMS LAB', '380':'OPERATIONS RESEARCH', '381':'BASIC ELECTRONICS', '382':'DIGITAL ELECTRONICS', '365':'ADVANCED COMPUTER SKILLS LAB', '402B':'OPERATING SYSTEMS', '402I':'OPERATING SYSTEM LAB', '410':'COMPUTER ORGANIZATION', '431':'SIGNALS AND SYSTEMS', '494':'MATHEMATICS-III', '472':'COMPUTER ORGANIZATION LAB', '473':'DATABASE MANAGEMENT SYS.LAB', '485':'EFFECTIVE TECH.COMM.IN ENGLISH', '488':'FINANCE AND ACCOUNTING', '448':'DATABASE MANAGEMENT SYSTEMS'}
valueTable_IT = {'120':'ENGLISH LANGUAGE LAB', '121':'C PROGRAMMING LAB','171':'MATHEMATICS-I', '172':'PHYSICS', '174':'BASIC ELECTRICAL ENGINEERING', '187':'ENGINEERING GRAPHICS', '186':'INDIAN CONSTITUTION', '179':'BASIC ELECTRICAL ENGINEERING LAB', '177':'PHYSICS LAB', '271':'ENGLISH', '272':'MATHEMATICS-II', '273A':'ENGINEERING CHEMISTRY', '274':'PROGRAMMING FOR PROBLEM SOLVING', '284':'WORKSHOP LAB', '278':'ESSENCE OF INDIAN TRADITIONAL KNOWLEDGE', '281':'ENGLISH LAB', '282A':'ENGINEERING CHEMISTRY LAB', '283':'PROGRAMMING FOR PROBLEM SOLVING LAB', '277':'ENVIRONMENTAL SCIENCE', '367':'BASIC ELECTRONICS LAB', '371':'DATA STRUCTURES LAB', '376':'IT WORKSHOP LAB', '381':'BASIC ELECTRONICS', '398':'MATHEMA.FOUN.OF INF.TECHNOLOGY', '391':'EFFECTIVE TECH.COMM.IN ENGLISH', '392':'FINANCE AND ACCOUNTING', '393':'MATHEMATICS-III', '397':'DATA STRUCTURES', '382':'DIGITAL ELECTRONICS', '398':'MATHEMA.FOUN.OF INF.TECHNOLOGY', '431':'SIGNALS AND SYSTEMS', '432':'COM.ORGAN.& MICROPROCESSOR', '498':'OPERATIONS RESEARCH', '449':'DATABASE SYSTEMS', '454':'JAVA PROGRAMMING LAB', '455':'MICROPROCESSOR LAB', '474':'DATABASE SYSTEMS LAB', '491':'JAVA PROGRAMMING', '434':'DATA COMMUNICATIONS'}
valueTable_GPA = {'S':10, 'A':9, 'B':8, 'C':7, 'D':6, 'E':5, 'F':0}

def extract_data(x):
    data = {}
    with open(x, 'r') as html_file:
        html_content = html_file.read()
        soup = BeautifulSoup(html_content, 'html.parser')
        gfg = soup.find(lambda tag: tag.name == "font" and "Name" in tag.text)
        v = list(gfg.parent.next_siblings)
        name = v[1].text
        gfg = soup.find_all(lambda tag: tag.name == "font" and "Grade Secured" in tag.text)
        for g in gfg:
            v = list(g.parent.parent.next_siblings)
            we_need = []
            for i in v:
                if(i != '\n'):
                    k = i.contents
                    for tag in k:
                        if tag != "\n":
                            we_need.append(tag)
            for i in range(0, len(we_need), 4):
                data[we_need[i].text.rstrip().lstrip().replace("\xa0", "")] = we_need[i+3].text.rstrip().lstrip().replace("\xa0", "")
                #print(data)
        gfg = soup.find(lambda tag: tag.name == "font" and "Result with SGPA" in tag.text)
        end_results = list(gfg.parent.parent.parent.next_siblings)
        #print(end_results)
        semester = {}
        for end in end_results:
            if end != '\n':
                s = end.text.replace('\n', '').replace('   ', ' ').strip().split("  ") #we have double space left
                semester[s[1]] = s[2]
    #print(name)
    #print(data)
    #print(semester)
    return name, data, semester


def xlw(l, s, valueTable):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    c = 1
    worksheet.cell(row=1, column=c).value = 'SNo'
    c += 1
    worksheet.cell(row=1, column=c).value = 'RNo'
    c += 1
    worksheet.cell(row=1, column=c).value = 'Name'
    c += 1
    for i in valueTable:
        worksheet.cell(row=1, column=c).value = i
        c += 1
        worksheet.merge_cells(start_row=1, start_column=c - 1, end_row=1, end_column=c)
        c += 1
    for i in range(1,9):
        worksheet.cell(row=1, column=c).value = i
        c += 1
    r = 2
    for i in range(0, len(l), 4):
        c = 1
        worksheet.cell(row=r, column=c).value = r - 1 #serial number
        c += 1
        worksheet.cell(row=r, column=c).value = l[i] #roll number
        c += 1
        worksheet.cell(row=r, column=c).value = l[i + 1] #name of the student
        c += 1
        while worksheet.cell(row=1, column=c).value != 1:
            try:
                worksheet.cell(row=r, column=c).value = valueTable_GPA[l[i + 2][worksheet.cell(row=1, column=c).value]]
                c += 1
                worksheet.cell(row=r, column=c).value = l[i + 2][worksheet.cell(row=1, column=c-1).value]
                c += 1
            except KeyError:
                worksheet.cell(row=r, column=c).value = 'NA'
                c += 1
                worksheet.cell(row=r, column=c).value = 'NA'
                c += 1
        while worksheet.cell(row=1, column=c).value != None:
            try:
                k = str(worksheet.cell(row=1, column=c).value)
                worksheet.cell(row=r, column=c).value = l[i + 3][k]
            except KeyError:
                worksheet.cell(row=r, column=c).value = 'NA'
            c += 1
        r+=1
    for i in range(4,c-9,2):
        worksheet.cell(row=1, column=i).value = valueTable[worksheet.cell(row=1, column=i).value]
    workbook.save(s)


global entry
global link


def getlink():
    link = entry.get()
    return link


def getrange():
    wn = tk.Tk()
    wn.geometry("750x250")
    global entry2
    global entry1

    def display_text():
        string1 = entry1.get()
        string2 = entry2.get()
        l = []
        link1=getlink()

        if(string1 <= string2):
            try:
                for i in range(int(string1), int(string2) + 1):
                    if i == 245621733041:
                        continue
                    else:
                        print(i)
                        payload = {'mbstatus': 'SEARCH', 'htno': i, 'Submit.x': 32, 'Submit.y': 6}
                        s = req.Session()
                        s.cert = "C:/Users/BSV/PycharmProjects/ou-results/www.osmania.ac.in.crt"

                        try:
                            resp = req.post(link1, data=payload, allow_redirects=True, verify=False)
                            f = open("info.html", "w")
                            f.write(resp.text)
                            f.close()
                            try:
                                lname, ldata, lcgpa = extract_data("info.html")
                                l.append(str(i))
                                l.append(lname)
                                l.append(ldata)
                                l.append(lcgpa)
                            except AttributeError:
                                #messagebox.showerror("showerror", "Page not found for "+str(i))
                                continue
                        except ConnectionError:
                            messagebox.showerror("showerror", "Connection failed")
                            break
                        # link error
                        except MissingSchema:
                            messagebox.showerror("showerror", "Incorrect URL. Please enter complete url in the format http://{url}")
                            break
                        except TimeoutError:
                            messagebox.showerror("showerror", "Connection timed out, please try again")
                            break

                        cert_reqs = 'CERT_REQUIRED'

                if string1[6:9] == '733':
                    xlw(l, "score_sheet.xlsx", valueTable_CSE)
                elif string1[6:9] == '737':
                    xlw(l, "score_sheet.xlsx", valueTable_IT)
                os.system("start EXCEL.EXE score_sheet.xlsx")
                # label.configure(text=string)
            except ValueError:
                messagebox.showerror("showerror", "Hallticket numbers seem incorrect")
        else:
            messagebox.showerror("showerror", "Hallticket range seems incorrect")

    tk.Label(wn, text="from", bg="white", fg='black', font=('Helvetica', 12)).place(x=100, y=70)
    tk.Label(wn, text="to", bg="white", fg='black', font=('Helvetica', 12)).place(x=100, y=100)
    # Initialize a Label to display the User Input
    label = Label(wn, text="", font=("Courier 22 bold"))
    label.pack()
    label1 = Label(wn, text="", font=("Courier 22 bold"))
    label1.pack()

    # Create an Entry widget to accept User Input
    entry1 = Entry(wn, width=20)
    entry1.focus_set()
    entry1.pack()

    entry2 = Entry(wn, width=20)
    entry2.focus_set()
    entry2.pack()

    # Create a Button to validate Entry Widget
    tk.Button(wn, text="Submit", bg='red', width=10, command=display_text).pack(pady=20)


def cse():
    l = []
    link2 = getlink()
    for i in range(245621733001, 245621733196):
        print(i)
        payload = {'mbstatus': 'SEARCH', 'htno': i, 'Submit.x': 32, 'Submit.y': 6}
        s = req.Session()
        s.cert = "C:/Users/BSV/PycharmProjects/ou-results/www.osmania.ac.in.crt"

        try:
            resp = req.post(link2, data=payload, allow_redirects=True, verify=False)
            f = open("info.html", "w")
            f.write(resp.text)
            f.close()
            try:
                cert_reqs = 'CERT_REQUIRED'
                lname, ldata, lcgpa = extract_data("info.html")
                l.append(str(i))
                l.append(lname)
                l.append(ldata)
                l.append(lcgpa)
            except AttributeError:
                #messagebox.showerror("showerror", "Page not found for " + str(i))
                continue
        except ConnectionError:
            messagebox.showerror("showerror", "Connection failed")
            break
        # link error
        except MissingSchema:
            messagebox.showerror("showerror", "Incorrect URL. Please enter complete url in the format http://{url}")
            break
        except TimeoutError:
            messagebox.showerror("showerror", "Connection timed out")
            break

    xlw(l, "score_sheet_cse.xlsx", valueTable_CSE)
    os.system("start EXCEL.EXE score_sheet.xlsx")



def it():
    l = []
    link3 = getlink()

    for i in range(245621737001, 245621737130):
        payload = {'mbstatus': 'SEARCH', 'htno': i, 'Submit.x': 32, 'Submit.y': 6}
        s = req.Session()
        s.cert = "C:/Users/BSV/PycharmProjects/ou-results/www.osmania.ac.in.crt"

        try:
            resp = req.post(link3, data=payload, allow_redirects=True, verify=False)
            f = open("info.html", "w")
            f.write(resp.text)
            f.close()
        except ConnectionError:
            messagebox.showerror("showerror", "Connection failed")
            break
        # link error
        except MissingSchema:
            messagebox.showerror("showerror", "Incorrect URL. Please enter complete url in the format http://{url}")
            break
        except TimeoutError:
            messagebox.showerror("showerror", "Connection timed out")
            break

        cert_reqs = 'CERT_REQUIRED'
        try:
            lname, ldata, lcgpa = extract_data("info.html")
            l.append(str(i))
            l.append(lname)
            l.append(ldata)
            l.append(lcgpa)
        except AttributeError:
            #messagebox.showerror("showerror", "Page not found for " + str(i))
            continue

    xlw(l, "score_sheet_it.xlsx", valueTable_IT)
    os.system("start EXCEL.EXE score_sheet.xlsx")


def RM():
    webbrowser.open_new_tab()


wn = tk.Tk()
wn.title("GLWEC Results")
wn.geometry('800x520')
wn.config(bg='antiquewhite')

tk.Label(wn, text="Exam Link: ", bg='antiquewhite', fg='black', font=('Helvetica', 12)).place(x=100, y=145)
entry = Entry(wn, width=50)
entry.focus_set()
entry.place(x=200, y=145)
tk.Label(wn, text='GLWEC RESULTS', bg='antiquewhite',
         fg='black', font=('Helvetica', 18, "bold")).place(x=250, y=75)
tk.Button(wn, text=" CSE ", bg='grey', font=('arial', 15), command=cse).place(x=190, y=200)
tk.Button(wn, text=" IT ", bg='grey', font=('arial', 15), command=it).place(x=300, y=200)
tk.Button(wn, text=" Get range ", bg='antiquewhite', font=('arial', 15), command=getrange).place(x=245, y=300)

showCommand = tk.StringVar()

wn.mainloop()